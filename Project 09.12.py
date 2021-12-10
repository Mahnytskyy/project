from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath




class NotAllData(Exception):
    pass

data_path=join('.', "table.xlsx")
data_path=abspath(data_path)

wb = load_workbook(filename="table.xlsx", data_only=True, read_only=True)

wsn = list(wb.sheetnames)
print(wsn)

wsdate = None
 
for i in wsn:
     if wb[i]['A2'].value == "Клієнт":
         wsdate = i
if wsdate == None:
    raise NotAllData('No data with Клієнт')

ws = wb[wsdate]
shapka = [cell.value for cell in next(ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))]

mandata ={}

for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
   if len(row) > 0:
    name_of_the_university = row[1].value
    if name_of_the_university is not None:
        name_of_the_universitydata =[cell.value for cell in row]
        if name_of_the_university not in mandata:
            mandata[name_of_the_university] = []
        mandata[name_of_the_university].append(name_of_the_universitydata)
        
        
for name_of_the_university in mandata:
    print(f'Клієнт {name_of_the_university}, Кількість продаж: {len(mandata[name_of_the_university])}')
    
    wb.close
    

for  name_of_the_university in mandata:
    exname, *_ =name_of_the_university.split()
    wb = Workbook()
    ws = wb.active
    ws.title = "Клієнт"
        
    ws.append(shapka)
    for row in mandata[name_of_the_university]:
        ws.append(row)
        
    for i in range(1,10):
        
        zagl = ws.cell(row=1, column=i)
        zagl.alignment = Alignment(horizontal='center')
        zagl.fill = PatternFill(fill_type='solid',start_color='5a61f0', end_color='5a61f0')
        zagl.font = Font(bold=True, italic=True,color='ffffff',size=18)
        
    nmrow = len(mandata[name_of_the_university])
    for i in range(2, nmrow + 2):
        ws.cell(row=i, column=1).number_format ='dd mmmmm yyyy'
        ws.cell(row=i, column=1).fill = PatternFill(
            fill_type='solid', start_color='4ae7f8', end_color='4ae7f8')
            
        ws.cell(row=i, column=2).fill = PatternFill(
            fill_type='solid', start_color='99ffcc', end_color='99ffcc')
            
        ws.cell(row=i, column=3).fill = PatternFill(
            fill_type='solid', start_color='4ae7f8', end_color='4ae7f8')
            
        ws.cell(row=i, column=4).number_format = '# ##0'
        ws.cell(row=i, column=4).fill = PatternFill(
            fill_type='solid', start_color='f8cd30', end_color='f8cd30')
            
        ws.cell(row=i, column=5).number_format = '# ##0.00'
        ws.cell(row=i, column=5).fill = PatternFill(
            fill_type='solid', start_color='f8cd30', end_color='f8cd30')
            
        ws.cell(row=i, column=6).number_format = '# ##0.00'
        ws.cell(row=i, column=6).fill = PatternFill(
            fill_type='solid', start_color='f8cd30', end_color='f8cd30')

        ws.cell(row=i, column=7).number_format = '# ##0.00%'
        ws.cell(row=i, column=7).fill = PatternFill(
            fill_type='solid', start_color='fda9fd', end_color='fda9fd')

        ws.cell(row=i, column=8).number_format = '# ##0.00'
        ws.cell(row=i, column=8).fill = PatternFill(
            fill_type='solid', start_color='37cd65', end_color='37cd65')
 
    ws.column_dimensions['A'].width =20
    ws.column_dimensions['B'].width =35
    ws.column_dimensions['C'].width =10
    ws.column_dimensions['D'].width =20
    ws.column_dimensions['E'].width =10
    ws.column_dimensions['F'].width =10
    ws.column_dimensions['G'].width =20
    ws.column_dimensions['H'].width =10
    ws.column_dimensions['I'].width =10
        
        
    exfilename = join('.', 'Data', (exname + '.xlsx'))
    exfilename = abspath(exfilename)
    print(exfilename)
        
    wb.save(exfilename)
    wb.close
    
print('Дані обработані')